#!/usr/bin/env python3
"""
Script to find devices inactive for more than 5 days, ping them, and generate a report
Output format: CSV/Excel file with Device ID, Token, Days Offline, Ping Status
"""

import sys
import os
import asyncio
import logging
from pathlib import Path
from datetime import datetime, timedelta, timezone
from typing import List, Dict, Any
import csv

# Local token validator to avoid heavy imports (e.g., fastapi) when running standalone
def is_valid_fcm_token(token) -> bool:
    if not token or not isinstance(token, str):
        return False
    token = token.strip()
    if not token or token.startswith("NO_FCM_TOKEN_"):
        return False
    return True

# Add parent directory to path
sys.path.insert(0, str(Path(__file__).parent.parent))

from app.config import settings
from app.database import connect_to_mongodb, mongodb, close_mongodb_connection
from app.utils.datetime_utils import utc_now
from app.services.device_service import DeviceService

# Import Firebase directly to avoid dependency issues
import firebase_admin
from firebase_admin import credentials, messaging
from pathlib import Path

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)


async def find_inactive_devices(days_threshold: int = 5) -> List[Dict[str, Any]]:
    """
    Find devices that haven't been pinged for more than specified days
    
    Args:
        days_threshold: Number of days to consider device inactive (default: 5)
    
    Returns:
        List of device documents
    """
    try:
        threshold_date = utc_now() - timedelta(days=days_threshold)
        
        logger.info(f"Searching for devices inactive since {threshold_date.isoformat()} (more than {days_threshold} days)")
        
        # Find devices where last_ping is older than threshold or doesn't exist
        # Exclude uninstalled devices
        query = {
            "is_deleted": {"$ne": True},
            "$and": [
                {
                    "$or": [
                        {"is_uninstalled": {"$ne": True}},
                        {"is_uninstalled": False},
                        {"is_uninstalled": {"$exists": False}}
                    ]
                },
                {
                    "$or": [
                        {"last_ping": {"$lt": threshold_date}},
                        {"last_ping": {"$exists": False}}
                    ]
                }
            ]
        }
        
        devices = await mongodb.db.devices.find(
            query,
            {
                "device_id": 1,
                "last_ping": 1,
                "fcm_tokens": 1,
                "model": 1,
                "manufacturer": 1,
                "registered_at": 1,
                "status": 1
            }
        ).to_list(length=None)
        
        logger.info(f"Found {len(devices)} inactive devices (more than {days_threshold} days)")
        
        return devices
        
    except Exception as e:
        logger.error(f"Error finding inactive devices: {e}", exc_info=True)
        return []


async def find_all_devices() -> List[Dict[str, Any]]:
    """
    Find ALL devices (not deleted)
    
    Returns:
        List of device documents
    """
    try:
        logger.info("Searching for ALL devices...")
        
        # Exclude uninstalled devices
        devices = await mongodb.db.devices.find(
            {
                "is_deleted": {"$ne": True},
                "$or": [
                    {"is_uninstalled": {"$ne": True}},
                    {"is_uninstalled": False},
                    {"is_uninstalled": {"$exists": False}}
                ]
            },
            {
                "device_id": 1,
                "last_ping": 1,
                "fcm_tokens": 1,
                "model": 1,
                "manufacturer": 1,
                "registered_at": 1,
                "status": 1
            }
        ).to_list(length=None)
        
        logger.info(f"Found {len(devices)} total devices (not deleted, not uninstalled)")
        
        return devices
        
    except Exception as e:
        logger.error(f"Error finding all devices: {e}", exc_info=True)
        return []


def calculate_days_offline(last_ping: datetime, current_time: datetime) -> float:
    """Calculate number of days since last ping"""
    if not last_ping:
        return None
    
    if isinstance(last_ping, str):
        try:
            last_ping = datetime.fromisoformat(last_ping.replace('Z', '+00:00'))
        except:
            return None
    
    if last_ping.tzinfo is None:
        last_ping = last_ping.replace(tzinfo=timezone.utc)
    
    delta = current_time - last_ping
    return round(delta.total_seconds() / 86400, 2)  # Convert to days


def is_valid_fcm_token(token: str) -> bool:
    """Check if FCM token is valid"""
    if not token or not isinstance(token, str):
        return False
    token = token.strip()
    if not token or token.startswith("NO_FCM_TOKEN"):
        return False
    return True


def init_firebase():
    """Initialize Firebase Admin SDK"""
    try:
        if not firebase_admin._apps:
            # Try to find Firebase credentials file
            base_dir = Path(__file__).parent.parent
            cred_files = ["apps.json", "admin.json"]
            cred_file = None
            
            for f in cred_files:
                path = base_dir / f
                if path.exists():
                    cred_file = str(path)
                    break
            
            if not cred_file:
                raise FileNotFoundError("Firebase credentials file not found (apps.json or admin.json)")
            
            cred = credentials.Certificate(cred_file)
            firebase_admin.initialize_app(cred)
            logger.info(f"Firebase initialized with {cred_file}")
        return True
    except Exception as e:
        logger.error(f"Firebase initialization error: {e}")
        return False


async def send_fcm_ping(token: str, device_id: str) -> bool:
    """Send FCM ping to a token"""
    try:
        message = messaging.Message(
            data={
                "type": "ping",
                "timestamp": str(int(utc_now().timestamp() * 1000))
            },
            token=token,
        )
        
        loop = asyncio.get_event_loop()
        response = await loop.run_in_executor(None, messaging.send, message)
        return True
    except messaging.UnregisteredError:
        logger.debug(f"Token unregistered for device {device_id}")
        return False
    except Exception as e:
        logger.debug(f"FCM error for device {device_id}: {e}")
        return False


async def ping_device_and_get_status(device_id: str) -> Dict[str, Any]:
    """
    Ping a device and return status
    
    Returns:
        Dict with ping result
    """
    try:
        # Get device from database
        device = await mongodb.db.devices.find_one(
            {"device_id": device_id},
            {"fcm_tokens": 1, "model": 1, "manufacturer": 1}
        )
        
        if not device:
            return {
                "success": False,
                "sent_count": 0,
                "failed_count": 0,
                "total_tokens": 0,
                "message": "Device not found"
            }
        
        tokens = device.get("fcm_tokens", [])
        if not tokens:
            return {
                "success": False,
                "sent_count": 0,
                "failed_count": 0,
                "total_tokens": 0,
                "message": "Device has no FCM tokens"
            }
        
        # Filter valid tokens
        valid_tokens = [token for token in tokens if is_valid_fcm_token(token)]
        
        if not valid_tokens:
            return {
                "success": False,
                "sent_count": 0,
                "failed_count": 0,
                "total_tokens": 0,
                "message": "Device has no valid FCM tokens"
            }
        
        # Send pings in parallel
        ping_tasks = [send_fcm_ping(token, device_id) for token in valid_tokens]
        results = await asyncio.gather(*ping_tasks, return_exceptions=True)
        
        success_count = sum(1 for r in results if r is True)
        failed_count = len(results) - success_count
        
        # If all pings failed, mark device as uninstalled
        if success_count == 0 and len(valid_tokens) > 0:
            logger.warning(f"⚠️ [PING] All {len(valid_tokens)} token(s) failed for device {device_id}. Marking as uninstalled.")
            try:
                await DeviceService.mark_device_uninstalled(device_id)
                logger.info(f"✅ [PING] Device {device_id} marked as uninstalled in database")
            except Exception as mark_error:
                logger.error(f"❌ [PING] Failed to mark device {device_id} as uninstalled: {mark_error}")
            
            return {
                "success": False,
                "sent_count": 0,
                "failed_count": failed_count,
                "total_tokens": len(valid_tokens),
                "is_uninstalled": True,
                "message": f"All {len(valid_tokens)} token(s) failed - Device marked as uninstalled"
            }
        
        return {
            "success": success_count > 0,
            "sent_count": success_count,
            "failed_count": failed_count,
            "total_tokens": len(valid_tokens),
            "message": f"Ping sent to {success_count}/{len(valid_tokens)} tokens"
        }
    except Exception as e:
        logger.error(f"Error pinging device {device_id}: {e}")
        return {
            "success": False,
            "sent_count": 0,
            "failed_count": 0,
            "total_tokens": 0,
            "message": str(e)
        }


async def process_single_device(device: Dict[str, Any], current_time: datetime) -> List[Dict[str, Any]]:
    """
    Process a single device: ping it and return results
    
    Returns:
        List of result entries (one per token)
    """
    results = []
    device_id = device.get("device_id")
    if not device_id:
        return results
    
    # Calculate days offline
    last_ping = device.get("last_ping")
    days_offline = calculate_days_offline(last_ping, current_time)
    
    # Get device info
    model = device.get("model", "Unknown")
    manufacturer = device.get("manufacturer", "Unknown")
    status = device.get("status", "unknown")
    fcm_tokens = device.get("fcm_tokens", [])
    
    # Filter valid tokens (local check to avoid heavy imports)
    valid_tokens = [token for token in fcm_tokens if is_valid_fcm_token(token)]
    
    # Ping device (async)
    ping_result = await ping_device_and_get_status(device_id)
    
    # Create result entry for each token
    if valid_tokens:
        for token in valid_tokens:
            results.append({
                "device_id": device_id,
                "token": token[:50] + "..." if len(token) > 50 else token,  # Truncate long tokens
                "full_token": token,
                "days_offline": days_offline if days_offline is not None else "N/A",
                "ping_status": "SUCCESS" if ping_result["success"] else "FAILED",
                "ping_sent": ping_result["sent_count"],
                "ping_failed": ping_result["failed_count"],
                "total_tokens": ping_result["total_tokens"],
                "ping_message": ping_result["message"],
                "model": model,
                "manufacturer": manufacturer,
                "device_status": status,
                "last_ping": last_ping.isoformat() if last_ping else "Never"
            })
    else:
        # No valid tokens
        results.append({
            "device_id": device_id,
            "token": "NO_VALID_TOKEN",
            "full_token": "",
            "days_offline": days_offline if days_offline is not None else "N/A",
            "ping_status": "NO_TOKEN",
            "ping_sent": 0,
            "ping_failed": 0,
            "total_tokens": 0,
            "ping_message": "Device has no valid FCM tokens",
            "model": model,
            "manufacturer": manufacturer,
            "device_status": status,
            "last_ping": last_ping.isoformat() if last_ping else "Never"
        })
    
    return results


async def process_devices(devices: List[Dict[str, Any]], days_threshold: int, concurrency: int = 50) -> List[Dict[str, Any]]:
    """
    Process devices in parallel: ping them and collect results
    
    Args:
        devices: List of device documents
        days_threshold: Days threshold (for logging)
        concurrency: Number of concurrent pings (default: 50)
    
    Returns:
        List of results with device info and ping status
    """
    results = []
    current_time = utc_now()
    
    total = len(devices)
    logger.info(f"Processing {total} devices in parallel (concurrency: {concurrency})...")
    
    # Process devices in batches for better control
    batch_size = concurrency
    
    for batch_start in range(0, total, batch_size):
        batch_end = min(batch_start + batch_size, total)
        batch = devices[batch_start:batch_end]
        batch_num = (batch_start // batch_size) + 1
        total_batches = (total + batch_size - 1) // batch_size
        
        logger.info(f"Processing batch {batch_num}/{total_batches} ({len(batch)} devices)...")
        
        # Create tasks for parallel execution
        tasks = [
            process_single_device(device, current_time)
            for device in batch
        ]
        
        # Execute all tasks in parallel
        batch_results = await asyncio.gather(*tasks, return_exceptions=True)
        
        # Collect results
        for idx, result in enumerate(batch_results):
            if isinstance(result, Exception):
                device_id = batch[idx].get("device_id", "unknown")
                logger.error(f"Error processing device {device_id}: {result}")
                # Add error entry
                results.append({
                    "device_id": device_id,
                    "token": "ERROR",
                    "full_token": "",
                    "days_offline": "N/A",
                    "ping_status": "ERROR",
                    "ping_sent": 0,
                    "ping_failed": 0,
                    "total_tokens": 0,
                    "ping_message": str(result),
                    "model": "Unknown",
                    "manufacturer": "Unknown",
                    "device_status": "unknown",
                    "last_ping": "Never"
                })
            else:
                results.extend(result)
        
        logger.info(f"Batch {batch_num} completed: {len(batch_results)} devices processed")
        
        # Small delay between batches to avoid overwhelming Firebase
        if batch_end < total:
            await asyncio.sleep(0.2)
    
    logger.info(f"All {total} devices processed. Total entries: {len(results)}")
    return results


def save_to_csv(results: List[Dict[str, Any]], output_file: str):
    """Save results to CSV file"""
    if not results:
        logger.warning("No results to save")
        return
    
    fieldnames = [
        "device_id",
        "token",
        "days_offline",
        "ping_status",
        "ping_sent",
        "ping_failed",
        "total_tokens",
        "ping_message",
        "model",
        "manufacturer",
        "device_status",
        "last_ping"
    ]
    
    with open(output_file, 'w', newline='', encoding='utf-8') as csvfile:
        writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
        writer.writeheader()
        
        for result in results:
            row = {field: result.get(field, "") for field in fieldnames}
            writer.writerow(row)
    
    logger.info(f"Results saved to CSV: {output_file}")


def save_to_excel(results: List[Dict[str, Any]], output_file: str, is_all_devices: bool = False):
    """Save results to Excel file (requires openpyxl)"""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
    except ImportError:
        logger.warning("openpyxl not installed. Install with: pip install openpyxl")
        logger.info("Falling back to CSV format")
        csv_file = output_file.replace('.xlsx', '.csv')
        save_to_csv(results, csv_file)
        return
    
    if not results:
        logger.warning("No results to save")
        return
    
    # Create workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    if is_all_devices:
        ws.title = "All Devices Ping Report"
    else:
        ws.title = "Inactive Devices Ping Report"
    
    # Headers
    headers = [
        "Device ID",
        "Token",
        "Days Offline",
        "Ping Status",
        "Ping Sent",
        "Ping Failed",
        "Total Tokens",
        "Ping Message",
        "Model",
        "Manufacturer",
        "Device Status",
        "Last Ping"
    ]
    
    # Write headers
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Write data
    field_mapping = {
        "Device ID": "device_id",
        "Token": "token",
        "Days Offline": "days_offline",
        "Ping Status": "ping_status",
        "Ping Sent": "ping_sent",
        "Ping Failed": "ping_failed",
        "Total Tokens": "total_tokens",
        "Ping Message": "ping_message",
        "Model": "model",
        "Manufacturer": "manufacturer",
        "Device Status": "device_status",
        "Last Ping": "last_ping"
    }
    
    for row_idx, result in enumerate(results, 2):
        for col_idx, header in enumerate(headers, 1):
            field = field_mapping[header]
            value = result.get(field, "")
            
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            
            # Color code ping status
            if header == "Ping Status":
                if value == "SUCCESS":
                    cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                elif value == "FAILED":
                    cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                elif value == "NO_TOKEN":
                    cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
            
            # Align numbers
            if header in ["Days Offline", "Ping Sent", "Ping Failed", "Total Tokens"]:
                cell.alignment = Alignment(horizontal="right", vertical="center")
            else:
                cell.alignment = Alignment(horizontal="left", vertical="center")
    
    # Auto-adjust column widths
    for col_idx, header in enumerate(headers, 1):
        max_length = len(header)
        for row_idx in range(2, len(results) + 2):
            cell_value = str(ws.cell(row=row_idx, column=col_idx).value or "")
            if len(cell_value) > max_length:
                max_length = len(cell_value)
        
        # Set column width (with some padding)
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_length + 2, 50)
    
    # Freeze first row
    ws.freeze_panes = "A2"
    
    # Calculate summary statistics
    unique_devices = set(r["device_id"] for r in results)
    devices_with_success = set(r["device_id"] for r in results if r["ping_status"] == "SUCCESS")
    devices_with_failed = set(r["device_id"] for r in results if r["ping_status"] == "FAILED")
    devices_with_no_token = set(r["device_id"] for r in results if r["ping_status"] == "NO_TOKEN")
    
    success_count = sum(1 for r in results if r["ping_status"] == "SUCCESS")
    failed_count = sum(1 for r in results if r["ping_status"] == "FAILED")
    no_token_count = sum(1 for r in results if r["ping_status"] == "NO_TOKEN")
    
    # Add summary at the bottom (2 rows gap, then summary)
    summary_start_row = len(results) + 4
    
    # Write summary headers
    summary_headers = [
        ["", ""],
        ["📊 SUMMARY STATISTICS", ""],
        ["", ""],
        ["Total Devices Processed", len(unique_devices)],
        ["Total Token Entries", len(results)],
        ["", ""],
        ["✅ Success Statistics", ""],
        ["Devices with successful ping", len(devices_with_success)],
        ["Successful token pings", success_count],
        ["", ""],
        ["❌ Failure Statistics", ""],
        ["Devices with failed ping", len(devices_with_failed)],
        ["Failed token pings", failed_count],
        ["", ""],
        ["⚠️ No Token Statistics", ""],
        ["Devices without valid tokens", len(devices_with_no_token)],
        ["No token entries", no_token_count],
        ["", ""],
        ["Note", "Each device may have multiple tokens. Each token = 1 row above."]
    ]
    
    summary_font_bold = Font(bold=True, size=11)
    summary_font_header = Font(bold=True, size=12, color="FFFFFF")
    summary_fill_header = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    summary_fill_section = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    
    for row_offset, row_data in enumerate(summary_headers):
        row_idx = summary_start_row + row_offset
        
        # Handle merged header row separately
        if row_offset == 1:  # Main header
            cell = ws.cell(row=row_idx, column=1, value=row_data[0])
            cell.fill = summary_fill_header
            cell.font = summary_font_header
            cell.alignment = Alignment(horizontal="left", vertical="center")
            # Merge cells for header (do this AFTER setting value)
            ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=len(headers))
        else:
            # Regular rows
            for col_idx, value in enumerate(row_data, 1):
                if col_idx <= len(row_data):
                    cell = ws.cell(row=row_idx, column=col_idx, value=value)
                    
                    # Style based on row type
                    if row_data[0] in ["✅ Success Statistics", "❌ Failure Statistics", "⚠️ No Token Statistics"]:
                        cell.fill = summary_fill_section
                        cell.font = summary_font_bold
                    elif row_data[1] == "" and row_data[0] != "" and row_data[0] != "Note":
                        cell.font = summary_font_bold
                    
                    cell.alignment = Alignment(horizontal="left", vertical="center")
    
    # Save file
    wb.save(output_file)
    logger.info(f"Results saved to Excel: {output_file}")
    logger.info(f"   - {len(unique_devices)} unique devices, {len(results)} total token entries")
    logger.info(f"   - Summary statistics added at the bottom of the sheet")


async def main():
    """Main function"""
    import argparse
    
    parser = argparse.ArgumentParser(
        description="Ping devices - All devices or inactive devices (more than N days)"
    )
    parser.add_argument(
        "--days",
        type=int,
        default=5,
        help="Number of days to consider device inactive (default: 5)"
    )
    parser.add_argument(
        "--output",
        type=str,
        default=None,
        help="Output file path (default: ping_report_YYYYMMDD_HHMMSS.xlsx)"
    )
    parser.add_argument(
        "--format",
        type=str,
        choices=["csv", "excel", "xlsx"],
        default="excel",
        help="Output format: csv or excel (default: excel)"
    )
    parser.add_argument(
        "--concurrency",
        type=int,
        default=50,
        help="Number of concurrent pings (default: 50, increase for faster processing)"
    )
    parser.add_argument(
        "--all-devices",
        action="store_true",
        help="Ping ALL devices (not just inactive ones)"
    )
    
    args = parser.parse_args()
    
    # Generate output filename if not provided
    if not args.output:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        if args.format == "csv":
            args.output = f"ping_report_{timestamp}.csv"
        else:
            args.output = f"ping_report_{timestamp}.xlsx"
    
    logger.info("=" * 60)
    if args.all_devices:
        logger.info("All Devices Ping Script")
    else:
        logger.info("Inactive Devices Ping Script")
    logger.info("=" * 60)
    if not args.all_devices:
        logger.info(f"Days threshold: {args.days}")
    logger.info(f"Output file: {args.output}")
    logger.info(f"Output format: {args.format}")
    logger.info(f"Concurrency: {args.concurrency}")
    logger.info("=" * 60)
    
    try:
        # Initialize Firebase
        logger.info("Initializing Firebase...")
        if not init_firebase():
            logger.error("Failed to initialize Firebase. Exiting.")
            return
        logger.info("✅ Firebase initialized")
        
        # Connect to MongoDB
        logger.info("Connecting to MongoDB...")
        await connect_to_mongodb()
        logger.info("✅ MongoDB connected")
        
        # Find devices
        if args.all_devices:
            logger.info(f"\nSearching for ALL devices...")
            devices = await find_all_devices()
        else:
            logger.info(f"\nSearching for devices inactive for more than {args.days} days...")
            devices = await find_inactive_devices(args.days)
        
        if not devices:
            logger.info("No devices found")
            return
        
        logger.info(f"Found {len(devices)} device(s)")
        
        # Process devices (ping them in parallel)
        logger.info(f"\nPinging devices in parallel (concurrency: {args.concurrency})...")
        start_time = datetime.now()
        results = await process_devices(devices, args.days, args.concurrency)
        elapsed_time = (datetime.now() - start_time).total_seconds()
        logger.info(f"Ping completed in {elapsed_time:.2f} seconds")
        
        logger.info(f"\nProcessed {len(results)} device/token combinations")
        
        # Save results
        logger.info(f"\nSaving results to {args.output}...")
        if args.format == "csv":
            save_to_csv(results, args.output)
        else:
            save_to_excel(results, args.output, is_all_devices=args.all_devices)
        
        # Print detailed summary
        logger.info("\n" + "=" * 60)
        logger.info("Detailed Summary")
        logger.info("=" * 60)
        
        success_count = sum(1 for r in results if r["ping_status"] == "SUCCESS")
        failed_count = sum(1 for r in results if r["ping_status"] == "FAILED")
        no_token_count = sum(1 for r in results if r["ping_status"] == "NO_TOKEN")
        
        # Count unique devices
        unique_devices = set(r["device_id"] for r in results)
        devices_with_success = set(r["device_id"] for r in results if r["ping_status"] == "SUCCESS")
        devices_with_failed = set(r["device_id"] for r in results if r["ping_status"] == "FAILED")
        devices_with_no_token = set(r["device_id"] for r in results if r["ping_status"] == "NO_TOKEN")
        
        logger.info(f"\n📊 Device Statistics:")
        logger.info(f"   Total devices processed: {len(devices)}")
        logger.info(f"   Unique devices in results: {len(unique_devices)}")
        logger.info(f"   Total token entries: {len(results)}")
        
        logger.info(f"\n✅ Success Statistics:")
        logger.info(f"   Devices with successful ping: {len(devices_with_success)}")
        logger.info(f"   Successful token pings: {success_count}")
        
        logger.info(f"\n❌ Failure Statistics:")
        logger.info(f"   Devices with failed ping: {len(devices_with_failed)}")
        logger.info(f"   Failed token pings: {failed_count}")
        
        logger.info(f"\n⚠️  No Token Statistics:")
        logger.info(f"   Devices without valid tokens: {len(devices_with_no_token)}")
        logger.info(f"   No token entries: {no_token_count}")
        
        # List devices without tokens
        if devices_with_no_token:
            logger.info(f"\n📋 Devices WITHOUT valid tokens ({len(devices_with_no_token)} devices):")
            for device_id in sorted(list(devices_with_no_token))[:20]:  # Show first 20
                logger.info(f"   - {device_id}")
            if len(devices_with_no_token) > 20:
                logger.info(f"   ... and {len(devices_with_no_token) - 20} more (see report file)")
        
        # List devices with failed pings
        if devices_with_failed:
            logger.info(f"\n📋 Devices with FAILED pings ({len(devices_with_failed)} devices):")
            failed_devices_list = sorted(list(devices_with_failed))[:20]  # Show first 20
            for device_id in failed_devices_list:
                # Get error message from results
                device_results = [r for r in results if r["device_id"] == device_id and r["ping_status"] == "FAILED"]
                if device_results:
                    error_msg = device_results[0].get("ping_message", "Unknown error")
                    logger.info(f"   - {device_id}: {error_msg[:60]}")
            if len(devices_with_failed) > 20:
                logger.info(f"   ... and {len(devices_with_failed) - 20} more (see report file)")
        
        logger.info(f"\n💾 Report saved to: {args.output}")
        logger.info("=" * 60)
        
    except Exception as e:
        logger.error(f"Error: {e}", exc_info=True)
    finally:
        await close_mongodb_connection()
        logger.info("MongoDB connection closed")


if __name__ == "__main__":
    asyncio.run(main())

